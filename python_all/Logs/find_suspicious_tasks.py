"""
Он ищет именно те признаки, которые указаны в рекомендациях для планировщика задач:
необычные задания, особенно от имени SYSTEM;
нечёткие имена задач;
запуск из %APPDATA%, Temp, Tmp, Downloads, Users\Public;
запуск через powershell, cmd /c, wscript, cscript, mshta, rundll32, regsvr32, installutil, certutil;
запуск скриптов .bat/.cmd/.ps1/.vbs/.js/.wsf/.hta;
UNC-пути \\server\share\...

Запуск:
python find_suspicious_tasks.py

По умолчанию он берёт все csv/xlsx/xlsm из папки work, расположенной рядом со скриптом.

Если папка другая:
python find_suspicious_tasks.py --work-dir /path/to/work

Что он создаёт:
work/out_tasks/suspicious_tasks_detailed.csv — детальные находки;
work/out_tasks/suspicious_tasks_summary_by_host.csv — сводка по хостам;
work/out_tasks/run_report.txt — краткий итог запуска.
 """

from __future__ import annotations

import csv
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Dict, Iterable, Iterator, List, Tuple

try:
    from openpyxl import load_workbook  # type: ignore
except Exception:  # pragma: no cover
    load_workbook = None


import argparse

BASE_DIR = Path(__file__).resolve().parent
DEFAULT_WORK_DIR = BASE_DIR / "work"


# При необходимости сюда можно добавить известные легитимные задачи/пути,
# которые нужно подавлять как ложноположительные.
ALLOWLIST_NAME_SUBSTRINGS: List[str] = []
ALLOWLIST_PATH_SUBSTRINGS: List[str] = []

SYSTEM_VALUES = {
    "система",
    "system",
    "nt authority\\system",
    "s-1-5-18",
}

SUSPICIOUS_NAME_PATTERNS: List[Tuple[re.Pattern[str], str, int]] = [
    (re.compile(r"\\?(myupdatetask|systemrecovery)$", re.I), "Нечёткое имя задачи из рекомендаций", 4),
    (re.compile(r"\\?(update(task)?|recovery|system(host|service)?|taskhost|hostservice|svchost|winupdate|windowsupdate)[-_ ]?\w*$", re.I), "Похожее на маскировку имя задачи", 3),
    (re.compile(r"\\?[a-z]{3,8}\d{2,}$", re.I), "Короткое нетипичное имя задачи с цифрами", 2),
]

SUSPICIOUS_INTERPRETER_PATTERNS: List[Tuple[re.Pattern[str], str, int]] = [
    (re.compile(r"powershell(\.exe)?\b", re.I), "Запуск PowerShell", 4),
    (re.compile(r"powershell(\.exe)?\b.*\s-(enc|encodedcommand)\b", re.I), "PowerShell с -enc/-EncodedCommand", 8),
    (re.compile(r"\bcmd(\.exe)?\b\s*/c\b", re.I), "Запуск cmd.exe /c", 3),
    (re.compile(r"\b(wscript|cscript|mshta|rundll32|regsvr32|installutil|certutil)(\.exe)?\b", re.I), "Запуск системного интерпретатора/LOLBIN", 4),
    (re.compile(r"\b(bat|cmd|ps1|vbs|js|jse|wsf|hta)\b", re.I), "Запуск скрипта", 3),
]

SUSPICIOUS_PATH_PATTERNS: List[Tuple[re.Pattern[str], str, int]] = [
    (re.compile(r"%appdata%|\\appdata\\", re.I), "Запуск из AppData", 5),
    (re.compile(r"%temp%|%tmp%|\\temp\\|\\tmp\\", re.I), "Запуск из Temp/Tmp", 6),
    (re.compile(r"\\downloads\\", re.I), "Запуск из Downloads", 5),
    (re.compile(r"\\users\\public\\", re.I), "Запуск из Users\\Public", 6),
    (re.compile(r"\\desktop\\", re.I), "Запуск с рабочего стола", 4),
    (re.compile(r"\\perflogs\\", re.I), "Запуск из PerfLogs", 5),
    (re.compile(r"^\\\\", re.I), "Запуск с UNC-пути", 3),
]

HEADER_ALIASES = {
    "host": ["Имя узла", "Hostname", "Host Name", "Host"],
    "task_name": ["Имя задачи", "TaskName", "Task Name", "Taskname"],
    "run_as": ["Запуск от имени", "Run As User", "Run As", "RunAsUser"],
    "action": ["Задача для выполнения", "Task To Run", "Action", "Actions", "Command"],
    "working_dir": ["Рабочая папка", "Start In", "Working Directory", "Рабочий каталог"],
    "author": ["Автор", "Author"],
    "logon_mode": ["Режим входа в систему", "Logon Mode"],
    "state": ["Состояние", "Status", "State"],
}


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in {"nan", "none", "н/д", "n/a"}:
        return ""
    return text


def pick(row: Dict[str, str], logical_name: str) -> str:
    for alias in HEADER_ALIASES[logical_name]:
        if alias in row:
            return normalize_text(row.get(alias, ""))
    return ""


def is_header_like(task_name: str, action: str, run_as: str, host: str) -> bool:
    values = {task_name.lower(), action.lower(), run_as.lower(), host.lower()}
    markers = {
        "имя задачи", "taskname", "task name",
        "задача для выполнения", "task to run",
        "запуск от имени", "run as user",
        "имя узла", "hostname", "host name",
    }
    return any(v in markers for v in values)


def open_csv_rows(path: Path) -> Iterator[Dict[str, str]]:
    encodings = ["utf-8-sig", "utf-8", "cp1251", "utf-16", "latin-1"]
    last_error = None
    for enc in encodings:
        try:
            with path.open("r", encoding=enc, newline="") as f:
                sample = f.read(4096)
                f.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=",;")
                except csv.Error:
                    dialect = csv.excel
                reader = csv.DictReader(f, dialect=dialect)
                for row in reader:
                    yield {str(k).strip(): normalize_text(v) for k, v in row.items() if k is not None}
            return
        except UnicodeError as e:
            last_error = e
            continue
    raise RuntimeError(f"Не удалось прочитать {path}: {last_error}")


def open_xlsx_rows(path: Path) -> Iterator[Dict[str, str]]:
    if load_workbook is None:
        raise RuntimeError("Для чтения XLSX требуется openpyxl")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = [normalize_text(v) for v in next(rows)]
    for values in rows:
        row = {headers[i]: normalize_text(values[i]) for i in range(min(len(headers), len(values)))}
        yield row


def iter_rows(path: Path) -> Iterator[Dict[str, str]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        yield from open_csv_rows(path)
    elif suffix in {".xlsx", ".xlsm"}:
        yield from open_xlsx_rows(path)
    else:
        return


def is_allowlisted(task_name: str, action: str, working_dir: str) -> bool:
    text_name = task_name.lower()
    text_path = f"{action} {working_dir}".lower()
    if any(x.lower() in text_name for x in ALLOWLIST_NAME_SUBSTRINGS):
        return True
    if any(x.lower() in text_path for x in ALLOWLIST_PATH_SUBSTRINGS):
        return True
    return False


def evaluate_row(row: Dict[str, str], source_file: str) -> Dict[str, str] | None:
    host = pick(row, "host")
    task_name = pick(row, "task_name")
    run_as = pick(row, "run_as")
    action = pick(row, "action")
    working_dir = pick(row, "working_dir")
    author = pick(row, "author")
    logon_mode = pick(row, "logon_mode")
    state = pick(row, "state")

    if is_header_like(task_name, action, run_as, host):
        return None

    if not any([host, task_name, run_as, action, working_dir, author, logon_mode, state]):
        return None

    if is_allowlisted(task_name, action, working_dir):
        return None

    score = 0
    reasons: List[str] = []

    run_as_norm = run_as.lower()
    if run_as_norm in SYSTEM_VALUES:
        reasons.append("Задание запускается от имени SYSTEM")
        score += 2

    for pattern, reason, weight in SUSPICIOUS_NAME_PATTERNS:
        if task_name and pattern.search(task_name):
            reasons.append(reason)
            score += weight
            break

    combined = " ".join(x for x in [action, working_dir] if x)

    for pattern, reason, weight in SUSPICIOUS_PATH_PATTERNS:
        if combined and pattern.search(combined):
            reasons.append(reason)
            score += weight

    for pattern, reason, weight in SUSPICIOUS_INTERPRETER_PATTERNS:
        if action and pattern.search(action):
            reasons.append(reason)
            score += weight

    if action and re.search(r"\.(bat|cmd|ps1|vbs|js|jse|wsf|hta)(\s|$|\")", action, re.I):
        reasons.append("В действии указан скриптовый файл")
        score += 3

    if not action:
        reasons.append("Нет команды запуска в выгрузке")
        score += 1

    if score == 0:
        return None

    if score >= 9:
        severity = "high"
    elif score >= 5:
        severity = "medium"
    else:
        severity = "low"

    return {
        "source_file": source_file,
        "host": host,
        "task_name": task_name,
        "run_as": run_as,
        "author": author,
        "logon_mode": logon_mode,
        "state": state,
        "action": action,
        "working_dir": working_dir,
        "severity": severity,
        "score": str(score),
        "reasons": "; ".join(dict.fromkeys(reasons)),
    }


def write_csv(path: Path, rows: List[Dict[str, str]]) -> None:
    if not rows:
        with path.open("w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow([
                "source_file", "host", "task_name", "run_as", "author", "logon_mode",
                "state", "action", "working_dir", "severity", "score", "reasons"
            ])
        return

    fieldnames = list(rows[0].keys())
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    parser = argparse.ArgumentParser(description="Поиск подозрительных заданий планировщика в файлах из папки work")
    parser.add_argument("--work-dir", default=str(DEFAULT_WORK_DIR), help="Путь к папке с CSV/XLSX файлами")
    args = parser.parse_args()

    work_dir = Path(args.work_dir).resolve()
    output_dir = work_dir / "out_tasks"
    output_dir.mkdir(parents=True, exist_ok=True)

    if not work_dir.exists():
        raise SystemExit(f"Папка не найдена: {work_dir}")

    input_files = sorted(
        [p for p in work_dir.iterdir() if p.is_file() and p.suffix.lower() in {".csv", ".xlsx", ".xlsm"}]
    )

    all_hits: List[Dict[str, str]] = []
    errors: List[Tuple[str, str]] = []

    for path in input_files:
        try:
            for row in iter_rows(path):
                hit = evaluate_row(row, path.name)
                if hit:
                    all_hits.append(hit)
        except Exception as e:
            errors.append((path.name, str(e)))

    all_hits.sort(key=lambda x: (-int(x["score"]), x["host"], x["task_name"], x["source_file"]))

    details_path = output_dir / "suspicious_tasks_detailed.csv"
    write_csv(details_path, all_hits)

    summary_rows: List[Dict[str, str]] = []
    by_host: Dict[str, List[Dict[str, str]]] = defaultdict(list)
    for hit in all_hits:
        by_host[hit["host"]].append(hit)

    for host, items in sorted(by_host.items(), key=lambda x: (-len(x[1]), x[0])):
        sev_counter = Counter(item["severity"] for item in items)
        reasons = Counter()
        for item in items:
            for reason in item["reasons"].split("; "):
                if reason:
                    reasons[reason] += 1
        top_reasons = "; ".join(f"{reason} ({count})" for reason, count in reasons.most_common(5))
        summary_rows.append({
            "host": host,
            "hits": str(len(items)),
            "high": str(sev_counter.get("high", 0)),
            "medium": str(sev_counter.get("medium", 0)),
            "low": str(sev_counter.get("low", 0)),
            "top_reasons": top_reasons,
        })

    summary_path = output_dir / "suspicious_tasks_summary_by_host.csv"
    write_csv(summary_path, summary_rows)

    report_lines = [
        f"Файлов обработано: {len(input_files)}",
        f"Подозрительных задач найдено: {len(all_hits)}",
        f"Хостов с находками: {len(by_host)}",
        f"Детальный отчёт: {details_path}",
        f"Сводка по хостам: {summary_path}",
    ]
    if errors:
        report_lines.append("")
        report_lines.append("Ошибки чтения файлов:")
        for name, err in errors:
            report_lines.append(f"- {name}: {err}")

    report_path = output_dir / "run_report.txt"
    report_path.write_text("\n".join(report_lines), encoding="utf-8")

    print("\n".join(report_lines))


if __name__ == "__main__":
    main()
